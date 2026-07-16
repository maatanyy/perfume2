"""플랫 테이블 엑셀 리포트 테스트"""

import openpyxl
from utils.excel_report import save_results_excel, coupon_rate_display

RESULTS = [
    {
        "product_id": 1,
        "product_name": "CK ALL 오 드 뚜알렛 100ml",
        "timestamp": "2026-07-04T12:00:00",
        "result_status": "success",
        "prices": [
            {"seller": "waffle", "상품 url": "http://waffle/1", "판매가": 45000,
             "쿠폰적용가": 40500, "상품 가격": 40500, "배송비": 2500,
             "배송비 여부": "유료", "최종 가격": 43000, "결과 상태": "success"},
            {"seller": "cj", "상품 url": "http://cj/1", "판매가": 43000,
             "쿠폰적용가": None, "상품 가격": 43000, "배송비": 0,
             "배송비 여부": "무료", "최종 가격": 43000, "결과 상태": "success"},
            {"seller": "gs", "상품 url": "http://gs/1", "판매가": None,
             "쿠폰적용가": None, "상품 가격": None, "배송비": None,
             "배송비 여부": "매진/품절", "최종 가격": None,
             "결과 상태": "sold_out", "에러 발생": "품절 감지"},
        ],
    },
    {
        "product_id": 2,
        "product_name": "딥디크 오로즈 EDT 50ml",
        "timestamp": "2026-07-04T12:01:00",
        "result_status": "success",
        "prices": [
            {"seller": "waffle", "상품 url": "http://waffle/2", "판매가": 98000,
             "쿠폰적용가": 93100, "상품 가격": 93100, "배송비": 0,
             "배송비 여부": "무료", "최종 가격": 93100, "결과 상태": "success"},
            {"seller": "lotte", "상품 url": "http://lotte/2", "판매가": 95000,
             "쿠폰적용가": 89000, "상품 가격": 89000, "배송비": 0,
             "배송비 여부": "무료", "최종 가격": 89000, "결과 상태": "success"},
        ],
    },
]

EXPECTED_HEADERS = ["제품명", "판매처", "URL", "판매가", "쿠폰율", "쿠폰적용가",
                    "배송비", "최종가격", "상태", "비고"]


def make_workbook(tmp_path):
    path = save_results_excel(RESULTS, "ssg", results_dir=str(tmp_path))
    assert path is not None
    return openpyxl.load_workbook(path)


def test_coupon_rate_display():
    assert coupon_rate_display(45000, 40500) == "10%"
    assert coupon_rate_display(45000, None) == "-"
    assert coupon_rate_display(None, 40500) == "-"
    assert coupon_rate_display(40000, 45000) == "-"   # 쿠폰가가 더 비싸면 무효


def test_flat_sheet_headers(tmp_path):
    wb = make_workbook(tmp_path)
    ws = wb["전체 결과"]
    headers = [c.value for c in ws[1]]
    assert headers == EXPECTED_HEADERS


def test_flat_sheet_row_count(tmp_path):
    wb = make_workbook(tmp_path)
    ws = wb["전체 결과"]
    # 헤더 1행 + 가격행 5개
    assert ws.max_row == 6


def test_flat_sheet_success_row(tmp_path):
    wb = make_workbook(tmp_path)
    ws = wb["전체 결과"]
    row = [c.value for c in ws[2]]  # 첫 데이터 행 = waffle/제품1
    assert row[0] == "CK ALL 오 드 뚜알렛 100ml"
    assert row[1] == "Waffle (우리회사)"
    assert row[3] == 45000       # 판매가 (숫자)
    assert row[4] == "10%"       # 쿠폰율
    assert row[5] == 40500       # 쿠폰적용가
    assert row[7] == 43000       # 최종가격
    assert row[8] == "판매중"


def test_flat_sheet_sold_out_row(tmp_path):
    wb = make_workbook(tmp_path)
    ws = wb["전체 결과"]
    row = [c.value for c in ws[4]]  # 제품1의 gs 행
    assert row[3] == "품절"      # 판매가 셀에 품절 명시
    assert row[7] == "-"
    assert row[8] == "품절"
    # 행 배경색 강조 확인 (빨간 계열)
    assert ws.cell(row=4, column=1).fill.start_color.rgb in ("FFFFC7CE", "00FFC7CE")


def test_reversal_sheet(tmp_path):
    wb = make_workbook(tmp_path)
    ws = wb["가격 역전 항목"]
    headers = [c.value for c in ws[1]]
    assert headers == EXPECTED_HEADERS + ["가격차이"]
    # 제품2만 역전 (lotte 89,000 < waffle 93,100) → waffle 행 + lotte 행
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    assert rows[0][1] == "Waffle (우리회사)"
    assert rows[1][1] == "경쟁사 (lotte)"
    assert rows[1][10] == "-4100원 저렴"


def test_reversal_sheet_empty(tmp_path):
    no_reversal = [RESULTS[0]]  # 제품1은 역전 없음
    path = save_results_excel(no_reversal, "ssg", results_dir=str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb["가격 역전 항목"]
    assert ws.cell(row=2, column=1).value == "가격 역전 항목이 없습니다."


def test_flat_sheet_missing_product_not_dropped(tmp_path):
    """product 레벨 크롤링 자체가 실패해 prices가 비어있으면
    제품 자체가 사라지지 않고 오류 행으로 남아야 한다."""
    results = RESULTS + [
        {
            "product_id": 3,
            "product_name": "실패한 제품 EDT 50ml",
            "timestamp": "2026-07-04T12:02:00",
            "result_status": "error",
            "prices": [],
            "error": "크롤링 실패: 모든 판매처에서 페이지 로드 불가",
        },
    ]
    path = save_results_excel(results, "ssg", results_dir=str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb["전체 결과"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    failed_rows = [r for r in rows if r[0] == "실패한 제품 EDT 50ml"]
    assert len(failed_rows) == 1
    row = failed_rows[0]
    assert row[1] == "-"   # 판매처
    assert row[2] == "-"   # URL
    assert row[3] == "-"   # 판매가
    assert row[8] == "오류"  # 상태
    assert "크롤링 실패" in row[9]  # 비고


def test_flat_sheet_missing_product_none_name(tmp_path):
    """product_name이 None이어도 'N/A'로 표시되어야 한다 (or 패턴 방어)."""
    results = [
        {
            "product_id": 4,
            "product_name": None,
            "timestamp": "2026-07-04T12:03:00",
            "result_status": "error",
            "prices": [],
        },
    ]
    path = save_results_excel(results, "ssg", results_dir=str(tmp_path))
    wb = openpyxl.load_workbook(path)
    ws = wb["전체 결과"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "N/A"
